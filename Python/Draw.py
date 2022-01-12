import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import matplotlib.colors as mcolors
from openpyxl import load_workbook
from openpyxl import Workbook
book=Workbook()
data=book.active
dfrow = load_workbook(filename="Python/Result1.xlsx",  data_only=True)
df = dfrow.active
def col(c,df):
    col=[]
    for i in range(2,31):
        col.append(df.cell(row=i,column=c).value)
    return col
n=col(1,df)
d=col(2,df)
M=col(3,df)
setuptime=col(4,df)
ppsize=col(5,df)
keygentime=col(6,df)
pksize=col(7,df)
secretkeysize=col(8,df)
Registrationtime=col(9,df)
Collateralsize=col(10,df)
Speendingtime=col(11,df)
ciphertext=col(12,df)
PPSpeendingtime=col(13,df)
Verificationtime=col(14,df)
PPVerificationtime=col(16,df)
decryptiontime=col(15,df)

fig, ((ax4, ax0, ax1),(ax2, ax3, ax5)) = plt.subplots(nrows=2, ncols=3,
                                    figsize=(14, 8))
#ax0.set_title('Registration time')
ax0.errorbar(n,Registrationtime,color='maroon', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6, label='Registration time')
ax0.grid()
ax0.legend(loc=2,prop={'size': 12})
ax0.set_xlabel('Number of Collaterals')
ax0.set_ylabel('Time (ms)')

#ax1.set_title('Collateral size')
ax1.plot(n,Collateralsize,color='maroon', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6,label='Collateral size')
ax1.grid()
ax1.legend(loc=2,prop={'size': 12})
ax1.set_xlabel('Number of Collaterals')
ax1.set_ylabel('Size (byte)')



#ax2.set_title('Speending time')
ax2.plot(d,Speendingtime,color='darkgreen', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='maroon', markersize=6,label='Spending time')
ax2.plot(d,PPSpeendingtime,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6,label='Pre-Processed' + '\n' + 'Spending time')
ax2.grid()
ax2.legend(loc=2,prop={'size': 12})
ax2.set_xlabel('Number of Collaterals')
ax2.set_ylabel('Time (sec)')


#ax3.set_title('ciphertext Size')
ax3.errorbar(n,ciphertext,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='purple', markersize=6,label='Ciphertext Size')
ax3.grid()
ax3.legend(loc=2,prop={'size': 12})
ax3.set_xlabel('Number of Collatorals')
ax3.set_ylabel('Size (kbyte)')
ax3.set_ylim([5.4,5.8])

#ax4.set_title('Key Gen time')
ax4.errorbar(M,keygentime,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6,label='Key Gen time')
ax4.grid()
ax4.legend(loc=2,prop={'size': 12})
ax4.set_xlabel('Number of Merchants')
ax4.set_ylabel('Time (ms)')
#ax4.set_ylim([0,7])

#ax5.set_title('Verification time')
ax5.errorbar(d,Verificationtime,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='g', markersize=6,label='Verification time')
ax5.grid()
ax5.legend(loc=2,prop={'size': 12})
ax5.set_xlabel('Number of Collaterals')
ax5.set_ylabel('Time (Sec)')
ax5.set_ylim([0.1,0.5])
plt.draw()
plt.savefig("performance.pdf", bbox_inches='tight')